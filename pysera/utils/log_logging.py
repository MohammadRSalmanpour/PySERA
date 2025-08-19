""" Functions for writing logging Errors and Warnings on excel file"""

import logging
from pathlib import Path
import os
import re
from typing import List, Tuple, Optional
from openpyxl import load_workbook, Workbook
import contextvars

# For multiprocessing-safe logging
import multiprocessing as mp
from logging.handlers import QueueHandler, QueueListener

logger = logging.getLogger("Dev_logger")

"""
Per-process logging context to inject image and ROI identifiers into log messages.
We use contextvars so it is safe across threads and processes. Each worker sets
the current image id once per image, and updates the current ROI name per ROI.
"""

# Context variables for current image and ROI
CURRENT_IMAGE_ID: contextvars.ContextVar[str] = contextvars.ContextVar("CURRENT_IMAGE_ID", default="")
CURRENT_ROI_NAME: contextvars.ContextVar[str] = contextvars.ContextVar("CURRENT_ROI_NAME", default="")


def set_image_context(image_id: str) -> None:
    try:
        CURRENT_IMAGE_ID.set(image_id or "")
    except Exception:
        pass


def set_roi_context(roi_name: Optional[str]) -> None:
    try:
        CURRENT_ROI_NAME.set(roi_name or "")
    except Exception:
        pass


def clear_roi_context() -> None:
    try:
        CURRENT_ROI_NAME.set("")
    except Exception:
        pass


class ContextInjectFilter(logging.Filter):
    """Filter that injects [image_id] and ROI 'name' into log messages when missing."""

    def filter(self, record: logging.LogRecord) -> bool:
        try:
            image_id = CURRENT_IMAGE_ID.get("")
            roi_name = CURRENT_ROI_NAME.get("")

            # Avoid double-formatting if the message already contains an image tag
            msg = str(record.getMessage())

            # Prepend [image_id] if available and not already present at start
            if image_id and not msg.lstrip().startswith("["):
                msg = f"[{image_id}] " + msg

            # Ensure ROI 'name' appears somewhere if we have it and it's not already present
            if roi_name and ("ROI '" not in msg):
                # Prefer to insert after leading [image] tag for readability
                if msg.startswith("["):
                    # Insert after the first closing bracket and a space
                    closing = msg.find("]")
                    if closing != -1:
                        msg = msg[:closing + 1] + f" ROI '{roi_name}' -" + msg[closing + 1:]
                    else:
                        msg = f"ROI '{roi_name}': " + msg
                else:
                    msg = f"ROI '{roi_name}': " + msg

            # Assign back to record.msg only if no args to avoid interfering with %-format args
            if not getattr(record, "args", None):
                record.msg = msg
                record.message = msg
        except Exception:
            # Never fail logging due to filter errors
            pass
        return True


class MemoryLogHandler(logging.Handler):
    def __init__(self):
        super().__init__()
        self.records = []

    def emit(self, record):
        if record.levelno >= self.level:
            self.records.append(self.format(record))

    def get_logs(self):
        return self.records

    def clear(self):
        self.records.clear()


def log_logger(excel_path, logs):
    try:
        # Convert path to Path object for robust handling
        excel_path = Path(excel_path)
        # Create parent directories if they don't exist
        excel_path.parent.mkdir(parents=True, exist_ok=True)

        # Always write logs (create workbook if missing)
        write_logs_to_excel(excel_path, logs, "Sheet_3")

    except ValueError as e:
        logger.error(f"Value error: {e}")
        raise
    except PermissionError as e:
        logger.error(f"Permission error: {e}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        raise


def setup_multiprocessing_logging(memory_handler: Optional[MemoryLogHandler]):
    """Create a multiprocessing Queue and a QueueListener that forwards child logs
    back to the main process. The listener writes to both console and the given
    memory_handler (for bug sheet capture).

    Returns (log_queue, queue_listener).
    """
    from pysera.config.settings import LOGGING_CONFIG  # local import to avoid cycles

    log_queue: mp.Queue = mp.Queue()

    # Console handler in main process
    console_handler = logging.StreamHandler()
    console_handler.setLevel(getattr(logging, LOGGING_CONFIG['console_level']))
    console_format = logging.Formatter(LOGGING_CONFIG['console_format'])
    console_handler.setFormatter(console_format)

    handlers = [console_handler]
    if memory_handler is not None:
        memory_handler.setLevel(getattr(logging, LOGGING_CONFIG['memory_level']))
        memory_format = logging.Formatter(LOGGING_CONFIG['console_format'])
        memory_handler.setFormatter(memory_format)
        handlers.append(memory_handler)

    listener = QueueListener(log_queue, *handlers, respect_handler_level=True)
    return log_queue, listener


def init_worker_logging(log_queue, console_level: str, console_format: str):
    """Initializer for worker processes to route logs to the main process via Queue."""
    logger = logging.getLogger('Dev_logger')
    logger.setLevel(logging.DEBUG)

    # Clear any inherited handlers to avoid duplicate output
    if logger.hasHandlers():
        logger.handlers.clear()

    # Queue handler sends records to the main process
    qh = QueueHandler(log_queue)
    qh.setLevel(getattr(logging, console_level))
    # Formatting is applied in the listener's handlers
    logger.addHandler(qh)
    logger.propagate = False

    # Add context injection filter so messages always include image/ROI when available
    logger.addFilter(ContextInjectFilter())


def _parse_level_and_message(log_line: str) -> Tuple[str, str]:
    """Extract LEVEL and MESSAGE from a formatted log line like:
    "2025-01-01 00:00:00 - WARNING - [PatientX] message..."
    Returns (level, message).
    """
    # Split only the first two " - " delimiters
    parts = log_line.split(" - ", maxsplit=2)
    if len(parts) == 3:
        _, level, message = parts
        return level.strip(), message.strip()
    # Fallback if unexpected format
    return "INFO", log_line.strip()


def _parse_patient_and_roi(message: str) -> Tuple[str, str, str]:
    """Extract PatientID and RoI from the message when present.
    Returns (patient_id, roi, clean_message).
    """
    patient_id = ""
    roi_name = ""
    clean_message = message

    # Pattern 1: Leading [PatientID]
    m = re.match(r"^\[(?P<pid>[^\]]+)\]\s*(?P<rest>.*)$", clean_message)
    if m:
        patient_id = m.group("pid").strip()
        clean_message = m.group("rest").strip()

    # Pattern 2: ROI 'name'
    m2 = re.search(r"ROI\s*'(?P<roi>[^']+)'", clean_message)
    if m2:
        roi_name = m2.group("roi").strip()

    # Pattern 3: File <name>:
    if not patient_id:
        m3 = re.search(r"\bFile\s+(?P<file>[^:]+):", clean_message)
        if m3:
            patient_id = m3.group("file").strip()

    return patient_id, roi_name, clean_message


def write_logs_to_excel(excel_path: str,
                        # RoI_name: str, add later
                        logs: Optional[List[str]],
                        sheet2write: str) -> None:
    BUG_SHEET_HEADERS = ["PatientID", "RoI", "Level", "Message"]

    # Prepare workbook and sheet
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet2write
        ws.append(BUG_SHEET_HEADERS)
    else:
        wb = load_workbook(excel_path)
        if sheet2write not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet2write)
            ws.append(BUG_SHEET_HEADERS)
        else:
            ws = wb[sheet2write]

    # Append parsed logs
    if logs:
        for log_line in logs:
            level, message = _parse_level_and_message(log_line)
            patient_id, roi_name, clean_message = _parse_patient_and_roi(message)
            ws.append([patient_id, roi_name, level, clean_message])

    wb.save(excel_path)
