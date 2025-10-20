""" Logging utilities for writing errors and warnings to Excel files. """

import logging
import multiprocessing as mp
import re
from logging.handlers import QueueListener, QueueHandler
from pathlib import Path
from typing import List, Tuple, Optional

from openpyxl import Workbook, load_workbook

from ..config.settings import LOGGING_CONFIG, LOG_LEVEL_MAP


# ------------------------------------------------------------
# Logging Filters and Handlers
# ------------------------------------------------------------

class MemoryLogHandler(logging.Handler):
    """Stores log records in memory for later use."""

    def __init__(self):
        super().__init__()
        self.records: List[str] = []

    def emit(self, record: logging.LogRecord) -> None:
        if record.levelno >= self.level:
            self.records.append(self.format(record))

    def get_logs(self) -> List[str]:
        return self.records.copy()

    def clear(self) -> None:
        self.records.clear()


# ------------------------------------------------------------
# Log Writing to Excel
# ------------------------------------------------------------

BUG_SHEET_HEADERS = ["PatientID", "RoI", "Level", "Message"]


def log_to_excel(excel_path: str, logs: List[str], sheet_name: str = "Report") -> None:
    """Write parsed logs into an Excel sheet."""
    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = create_or_load_workbook(excel_path)
    worksheet = ensure_sheet_exists(workbook, sheet_name, BUG_SHEET_HEADERS)

    append_logs_to_sheet(worksheet, logs)
    workbook.save(excel_path)


def create_or_load_workbook(path: Path) -> Workbook:
    """Load an existing workbook or create a new one."""
    return load_workbook(path) if path.exists() else Workbook()


def ensure_sheet_exists(workbook: Workbook, sheet_name: str, headers: List[str]):
    """Ensure a sheet exists and has headers."""
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(headers)
    else:
        sheet = workbook[sheet_name]
    return sheet


def append_logs_to_sheet(sheet, logs: List[str]) -> None:
    """Append parsed log entries to the Excel sheet."""
    for line in logs:
        level, message = parse_log_level_and_message(line)
        patient_id, roi_name, clean_message = parse_patient_and_roi(message)
        sheet.append([patient_id, roi_name, level, clean_message])


# ------------------------------------------------------------
# Log Parsing
# ------------------------------------------------------------

def parse_log_level_and_message(log_line: str) -> Tuple[str, str]:
    """Extract log level and message from formatted log line."""
    parts = log_line.split(" - ", maxsplit=2)
    if len(parts) == 3:
        _, level, message = parts
        return level.strip(), message.strip()
    return "INFO", log_line.strip()


def parse_patient_and_roi(message: str) -> Tuple[str, str, str]:
    """Extract patient ID and ROI name from a message."""
    patient_id = extract_patient_id(message)
    roi_name = extract_roi_name(message)
    clean_message = clean_message_text(message)
    return patient_id, roi_name, clean_message


def extract_patient_id(message: str) -> str:
    """Extract patient ID from [PatientID] or File <name>: pattern."""
    match = re.match(r"^\[(?P<pid>[^\]]+)\]", message)
    if match:
        return match.group("pid").strip()
    file_match = re.search(r"\bFile\s+(?P<file>[^:]+):", message)
    return file_match.group("file").strip() if file_match else ""


def extract_roi_name(message: str) -> str:
    """Extract ROI name from message."""
    match = re.search(r"ROI\s*'(?P<roi>[^']+)'", message)
    return match.group("roi").strip() if match else ""


def clean_message_text(message: str) -> str:
    """Return cleaned message text."""
    return message.strip()


# ------------------------------------------------------------
# Utility Logging
# ------------------------------------------------------------

def get_levels_from_mode(log_mode_level: str = "all") -> Tuple[Optional[int], Optional[int]]:
    """Return console and memory logging levels based on LOG_LEVEL_MODE."""
    mode_config = LOG_LEVEL_MAP.get(log_mode_level, LOG_LEVEL_MAP["all"])
    console_level = getattr(logging, mode_config['console_level']) if mode_config['console_level'] else None
    memory_level = getattr(logging, mode_config['memory_level']) if mode_config['memory_level'] else None
    return console_level, memory_level


def create_console_handler(console_level: Optional[int], log_level_mode: str = "all") -> Optional[
    logging.StreamHandler]:
    """Create a console handler if console_level is set."""
    if console_level is None:
        return None

    handler = logging.StreamHandler()
    handler.setLevel(console_level)

    if log_level_mode == "info":
        handler.addFilter(lambda record: record.levelno == logging.INFO)

    handler.setFormatter(logging.Formatter(LOGGING_CONFIG['console_format']))
    return handler


def configure_memory_handler(memory_handler: MemoryLogHandler, memory_level: Optional[int],
                             log_model_level: str = "all") -> None:
    """Configure memory handler if memory_level is set."""
    if memory_level is None:
        return
    memory_handler.setLevel(memory_level)

    # level of INFO = 20
    if log_model_level == "info":
        memory_handler.addFilter(lambda record: record.levelno == logging.INFO)

    memory_handler.setFormatter(logging.Formatter(LOGGING_CONFIG['memory_format']))


def setup_logging(memory_handler: Optional[MemoryLogHandler] = None, log_model_level: str = "all") -> Tuple[
    logging.Logger, Optional[MemoryLogHandler]]:
    """
    Setup standard logging with console and optional memory handler.
    Logging levels are determined by LOG_LEVEL_MODE.
    """
    logger = logging.getLogger("Dev_logger")
    logger.setLevel(logging.DEBUG)
    if logger.hasHandlers():
        logger.handlers.clear()

    console_level, memory_level = get_levels_from_mode(log_model_level)

    # Silent mode
    if console_level is None and memory_level is None:
        logger.disabled = True
        return logger, memory_handler

    # Console
    console_handler = create_console_handler(console_level, log_model_level)
    if console_handler:
        logger.addHandler(console_handler)

    # Memory
    if memory_handler:
        configure_memory_handler(memory_handler, memory_level, log_model_level)
        if memory_level is not None:
            logger.addHandler(memory_handler)

    return logger, memory_handler


def initialize_logging(log_mode_level: str = "all") -> Tuple[logging.Logger, MemoryLogHandler]:
    """Initialize logger and memory handler."""
    memory_handler = MemoryLogHandler()

    if log_mode_level == "info":
        memory_handler.addFilter(lambda record: record.levelno == logging.INFO)

    _logger, memory_handler = setup_logging(memory_handler, log_mode_level)
    return _logger, memory_handler


def setup_multiprocessing_logging(memory_handler: Optional[MemoryLogHandler] = None, log_mode_level: str = "all") -> \
Tuple[
    Optional[mp.Queue], Optional[QueueListener]]:
    """
    Set up multiprocessing logging using a QueueListener.
    Respects LOG_LEVEL_MODE.
    """
    console_level, memory_level = get_levels_from_mode(log_mode_level)

    # Silent mode
    if console_level is None and memory_level is None:
        logging.getLogger("Dev_logger").disabled = True
        return None, None

    log_queue = mp.Queue()
    handlers = []

    console_handler = create_console_handler(console_level, log_mode_level)
    if console_handler:
        handlers.append(console_handler)

    if memory_handler:
        configure_memory_handler(memory_handler, memory_level, log_mode_level)
        if memory_level is not None:
            handlers.append(memory_handler)

    listener = QueueListener(log_queue, *handlers, respect_handler_level=True)
    return log_queue, listener


def init_worker_logging(log_queue: mp.Queue, log_level_mode: str = "all") -> None:
    """
    Initialize worker logger to send logs to main process via queue.
    Respects LOG_LEVEL_MODE.
    """
    console_level, _ = get_levels_from_mode(log_level_mode)
    if console_level is None:
        logging.getLogger("Dev_logger").disabled = True
        return

    worker_logger = logging.getLogger("Dev_logger")
    worker_logger.setLevel(logging.DEBUG)
    worker_logger.handlers.clear()

    queue_handler = QueueHandler(log_queue)
    queue_handler.setLevel(console_level)
    worker_logger.addHandler(queue_handler)
    worker_logger.propagate = False
