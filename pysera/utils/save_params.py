"""
Functions for writing parameters on Excel file
"""

from pathlib import Path
from typing import Dict, Union
from openpyxl import Workbook, load_workbook


def write_to_excel(
        excel_path: Union[str, Path],
        params_data: Dict
) -> None:
    excel_path = Path(excel_path)
    params_dict = params_data

    # Ensure parent folder exists
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    # Load or create workbook
    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()

    # Ensure 'parameters' sheet exists
    if "Parameters" not in wb.sheetnames:
        ws = wb.create_sheet("Parameters")
    else:
        ws = wb["Parameters"]

    # Clear existing content
    ws.delete_rows(1, ws.max_row)

    # Write header (keys)
    ws.append(list(params_dict.keys()))
    # Write values
    ws.append(list(params_dict.values()))

    # Save workbook
    wb.save(excel_path)
